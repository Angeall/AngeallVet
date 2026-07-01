"""Build .xlsx workbooks and stream them as downloads (openpyxl).

A *sheet* is ``{"title": str, "headers": [str], "rows": [[scalar, ...]]}``.
Used by the export endpoints (generic stats export + full DB backup).
"""

import io
import re
import enum
from datetime import datetime, date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from fastapi.responses import StreamingResponse

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_BAD_TITLE = re.compile(r"[\[\]:*?/\\]")
_FORMULA_PREFIX = ("=", "+", "-", "@")


def _safe_title(title, used):
    """Excel sheet titles: <=31 chars, no []:*?/\\, unique within a workbook."""
    t = (_BAD_TITLE.sub(" ", str(title or "Feuille")).strip() or "Feuille")[:31]
    base, i = t, 1
    while t.lower() in used:
        suffix = f"_{i}"
        t = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(t.lower())
    return t


def csv_safe(v):
    """Neutralise CSV/spreadsheet formula injection for a raw CSV cell.

    A leading =,+,-,@ makes Excel/LibreOffice treat the cell as a formula; a
    tab/CR/LF would corrupt the row/column layout. Used by the CSV exports that
    don't go through the workbook builder (which already calls ``_cell``).
    """
    if v is None:
        return ""
    s = v if isinstance(v, str) else str(v)
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    if s and s[0] in _FORMULA_PREFIX:
        s = " " + s
    return s


def _cell(v):
    """Coerce a value to an Excel-safe scalar, neutralising formula injection."""
    if v is None or isinstance(v, bool) or isinstance(v, (int, float)):
        return v
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, enum.Enum):
        v = v.value
    s = v if isinstance(v, str) else str(v)
    # A leading =,+,-,@ would be parsed as a formula by Excel -> neutralise it.
    if s and s[0] in _FORMULA_PREFIX:
        return " " + s
    return s


def build_workbook(sheets):
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for s in sheets:
        ws = wb.create_sheet(title=_safe_title(s.get("title"), used))
        headers = s.get("headers") or []
        if headers:
            ws.append([_cell(h) for h in headers])
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
        for row in s.get("rows") or []:
            ws.append([_cell(c) for c in row])
    if not wb.sheetnames:
        wb.create_sheet(title="Vide")
    return wb


def workbook_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9_.-]", "_", filename) or "export.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{safe}"'},
    )
